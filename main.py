# This is a sample Python script.

# Press Shift+F10 to execute it or replace it with your code.
# Press Double Shift to search everywhere for classes, files, tool windows, actions, and settings.


import bs4
import requests
import pandas
import openpyxl
import datetime
from os.path import exists

FilePath = ("final_company.xlsx")
final_column_len = 0


def write_more_info(old_names,name_list, address_list, others_info):
    # 讀取 Excel 檔案
    wb = openpyxl.load_workbook(FilePath)
    sheet = wb.active
    final_column_len = len(old_names)

    x = 0
    if final_column_len > 0:
        #print(final_column_len)
        l = len(others_info)
        #print(final_column_len + l)
        for j in range(final_column_len+1, final_column_len+l):
            sheet.cell(row=j, column=1).value = name_list[x]
            sheet.cell(row=j, column=2).value = address_list[x]
            sheet.cell(row=j, column=3).value = others_info[x][1]
            sheet.cell(row=j, column=4).value = others_info[x][3]
            sheet.cell(row=j, column=5).value = others_info[x][5]
            sheet.cell(row=j, column=6).value = others_info[x][7]
            sheet.cell(row=j, column=7).value = others_info[x][9]
            sheet.cell(row=j, column=8).value = others_info[x][14]
            x = x + 1
    # print(others_info[0])
    wb.save(FilePath)
    wb.close()


def write_others_info(others_info):
    # 讀取 Excel 檔案
    wb = openpyxl.load_workbook(FilePath)
    sheet = wb.active
    # 以行號、列號指定儲存格
    sheet.cell(row=1, column=1).value = "公司名稱"
    sheet.cell(row=1, column=2).value = "地址"
    sheet.cell(row=1, column=3).value = others_info[0][0]
    sheet.cell(row=1, column=4).value = others_info[0][2]
    sheet.cell(row=1, column=5).value = others_info[0][4]
    sheet.cell(row=1, column=6).value = others_info[0][6]
    sheet.cell(row=1, column=7).value = others_info[0][8]
    sheet.cell(row=1, column=8).value = others_info[0][13]

    for j in range(2, len(others_info) + 2):
        sheet.cell(row=j, column=3).value = others_info[j - 2][1]
        sheet.cell(row=j, column=4).value = others_info[j - 2][3]
        sheet.cell(row=j, column=5).value = others_info[j - 2][5]
        sheet.cell(row=j, column=6).value = others_info[j - 2][7]
        sheet.cell(row=j, column=7).value = others_info[j - 2][9]
        sheet.cell(row=j, column=8).value = others_info[j - 2][14]

    #print(others_info[0])
    wb.save(FilePath)
    wb.close()


def get_company_info(xurl):
    tmp_str = []
    htmlfile2 = requests.get(xurl)
    # print(htmlfile.status_code)
    # print(htmlfile2.text)
    # tmp_str = str(htmlfile2.text).split("profile")[1]
    # print(tmp_str)

    objssoup = bs4.BeautifulSoup(htmlfile2.text, 'lxml')
    objstag = objssoup.select('body div div div div div div div div div div')

    for i in range(0, len(objstag)):

        if "公司網址" and "產業類別" in objstag[i].text:

            if len(objstag[i].text) > 80:
                # print("=========================== start   ========================at" + str(i))
                # print(objstag[i].text)
                objstag[i].text  # .split("產業類別")[1].split("產業描述")[0]
                tmp_str.append(objstag[i].text)

    return tmp_str


def print_hi():
    file_exists = exists(FilePath)
    max_col = 0
    old_datas = []
    if file_exists is True:
        ExcelWorkbook = openpyxl.load_workbook(FilePath)
        writer = pandas.ExcelWriter(FilePath, engine='openpyxl')
        writer.book = ExcelWorkbook

        ws = ExcelWorkbook.active
        first_column = ws['A']
        final_column_len = len(first_column)
        # Print the contents
        for x in range(len(first_column)):
            old_datas.append(first_column[x].value)
        # print(old_datas)

    tmp_name_target = []
    tmp_address = []
    tmp_others = []
    tmp_others_list = []
    for j in range(1, 50):
        print("page" + str(j))
        # keyword: NX UG
        turl = "https://www.104.com.tw/jobs/search/?keyword=NX%20UG&order=1&jobsource=2018indexpoc&ro=0&page=" + str(j)
        htmlfile = requests.get(turl)
        # print(htmlfile.status_code)
        # print(htmlfile.text)
        if "搜尋條件無符合工作機會，建議放寬條件重新查詢" in htmlfile.text:
            # print(j)
            break
        else:
            objssoup = bs4.BeautifulSoup(htmlfile.text, 'lxml')
            objstag = objssoup.select('ul li a')
            # print(objstag)
            # print(str(objstag[5]).split("公司住址：")[1].split('">')[0])
            # print(str(objstag[5]).split('" target=')[0].split('href="')[1])

            for i in range(5, len(objstag)):
                if objstag[i].text.rstrip() not in tmp_name_target:
                    if objstag[i].text.rstrip() not in old_datas:
                        tmp_name_target.append(objstag[i].text.rstrip())
                        tmp_address.append(str(objstag[i]).split("公司住址：")[1].split('">')[0])
                        tmp_others = get_company_info(
                            "https:" + str(objstag[i]).split('" target=')[0].split('href="')[1])
                        tmp_others_list.append(tmp_others[0].split("  "))

    if file_exists is True:
        writer.close()
        final_column_len = len(tmp_name_target)
        #print(final_column_len)
        if final_column_len > 0:
            print("new_company："+ str(final_column_len-1))
            write_more_info(old_datas, tmp_name_target, tmp_address,  tmp_others_list)


    else:
        final_target = pandas.Series(tmp_address, index=tmp_name_target)
        final_target.to_excel(FilePath, header=False, startcol=0, startrow=1, sheet_name='Sheet1')
        write_others_info(tmp_others_list)


# Press the green button in the gutter to run the script.
if __name__ == '__main__':
    print_hi()

# See PyCharm help at https://www.jetbrains.com/help/pycharm/